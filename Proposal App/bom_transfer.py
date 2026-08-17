"""
Reads values from BOM Excel and writes them into the Proposal Input Excel.
Preserves destination cell formatting — only sets .value, never touches .number_format or style.
"""

import openpyxl


def transfer_values(bom_path, proposal_path, output_path, mappings):
    """
    Transfer BOM values into Proposal Input Excel and save to output_path.
    Returns (success: bool, results: list[dict], warnings: list[str]).
    Missing sheets or empty cells produce warnings and are skipped — not hard errors.
    """
    bom_wb = openpyxl.load_workbook(bom_path, data_only=True)
    proposal_wb = openpyxl.load_workbook(proposal_path)

    results = []
    warnings = []

    for m in mappings:
        src_sheet = m["source_sheet"]
        src_cell = m["source_cell"]
        tgt_sheet = m["target_sheet"]
        tgt_cell = m["target_cell"]
        label = m.get("label", "")

        # Check source sheet exists
        if src_sheet not in bom_wb.sheetnames:
            warnings.append(
                f"⚠ Skipped: BOM sheet '{src_sheet}' not found ({label})"
            )
            continue

        # Check target sheet exists
        if tgt_sheet not in proposal_wb.sheetnames:
            warnings.append(
                f"⚠ Skipped: Proposal Input sheet '{tgt_sheet}' not found ({label})"
            )
            continue

        # Check source cell has a value
        src_val = bom_wb[src_sheet][src_cell].value
        if src_val is None:
            warnings.append(
                f"⚠ Skipped: {src_sheet}!{src_cell} is empty ({label})"
            )
            continue

        # Transfer value — formatting untouched
        proposal_wb[tgt_sheet][tgt_cell].value = src_val
        results.append({
            "source": f"{src_sheet}!{src_cell}",
            "target": f"{tgt_sheet}!{tgt_cell}",
            "value": src_val,
            "label": label,
        })

    proposal_wb.save(output_path)
    return True, results, warnings
